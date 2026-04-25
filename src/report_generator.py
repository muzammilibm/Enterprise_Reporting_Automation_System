"""
Report Generator Module

This module handles the generation of Excel reports with formatted data.
It creates professional Excel files with multiple sheets including daily summaries,
month-to-date summaries, and validation warnings.

Author: Intern Project - Day 3
Date: 2026-04-25
"""

from datetime import datetime
from pathlib import Path
from typing import Tuple, Optional
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def generate_excel_report(
    daily_metrics: dict,
    mtd_metrics: dict,
    output_dir: str,
    validation_warnings: list
) -> Tuple[bool, Optional[str], Optional[str]]:
    """
    Generate an Excel report with daily and MTD metrics.
    
    Creates a timestamped Excel file with multiple sheets:
    - Sheet 1: Daily Summary with current day's metrics
    - Sheet 2: MTD Summary with month-to-date metrics
    - Sheet 3: Warnings (only if validation_warnings is not empty)
    
    Args:
        daily_metrics (dict): Dictionary containing daily metrics with keys:
            - total_count: Total number of records
            - total_processed: Number of successfully processed records
            - total_errors: Number of error records
            - processing_rate: Processing success rate as percentage
            - error_rate: Error rate as percentage
        mtd_metrics (dict): Dictionary containing MTD metrics with keys:
            - mtd_total: Total records for the month
            - mtd_processed: Total processed records for the month
            - mtd_errors: Total error records for the month
            - processing_rate: MTD processing rate as percentage
            - error_rate: MTD error rate as percentage
            - num_days: Number of days in the period
            - daily_average: Average records per day
        output_dir (str): Directory path where the report will be saved
        validation_warnings (list): List of warning messages to include
    
    Returns:
        Tuple[bool, Optional[str], Optional[str]]: 
            - success: True if report was generated successfully, False otherwise
            - file_path: Full path to the generated file, or None if failed
            - error_message: Error description if failed, or None if successful
    
    Example:
        >>> daily = {'total_count': 2850, 'total_processed': 2707, ...}
        >>> mtd = {'mtd_total': 42750, 'mtd_processed': 40605, ...}
        >>> success, path, error = generate_excel_report(daily, mtd, "./output", [])
        >>> if success:
        ...     print(f"Report saved: {path}")
    """
    try:
        print("Creating Excel file...")
        
        # Generate timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.xlsx"
        
        # Ensure output directory exists
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Full file path
        file_path = output_path / filename
        
        # Create new workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet and create our sheets
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)
        
        # Create Sheet 1: Daily Summary
        print("Creating Daily Summary sheet...")
        daily_sheet = workbook.create_sheet("Daily Summary", 0)
        _create_daily_summary_sheet(daily_sheet, daily_metrics)
        
        # Create Sheet 2: MTD Summary
        print("Creating MTD Summary sheet...")
        mtd_sheet = workbook.create_sheet("MTD Summary", 1)
        _create_mtd_summary_sheet(mtd_sheet, mtd_metrics)
        
        # Create Sheet 3: Warnings (only if there are warnings)
        if validation_warnings:
            print(f"Creating Warnings sheet with {len(validation_warnings)} warning(s)...")
            warnings_sheet = workbook.create_sheet("Warnings", 2)
            _create_warnings_sheet(warnings_sheet, validation_warnings)
        
        # Save the workbook
        print(f"Saving report to: {file_path}")
        workbook.save(file_path)
        
        print(f"[OK] Report successfully created: {file_path}")
        return True, str(file_path), None
        
    except PermissionError as e:
        error_msg = f"Permission denied: Cannot write to {output_dir}. Check file permissions."
        print(f"[X] Error: {error_msg}")
        return False, None, error_msg
        
    except OSError as e:
        error_msg = f"OS error: {str(e)}. Check disk space and path validity."
        print(f"[X] Error: {error_msg}")
        return False, None, error_msg
        
    except Exception as e:
        error_msg = f"Unexpected error generating report: {str(e)}"
        print(f"[X] Error: {error_msg}")
        return False, None, error_msg


def _create_daily_summary_sheet(sheet, daily_metrics: dict) -> None:
    """
    Create and populate the Daily Summary sheet.
    
    Args:
        sheet: openpyxl worksheet object
        daily_metrics (dict): Dictionary containing daily metrics
    """
    # Define headers
    headers = [
        "Date",
        "Total Count",
        "Processed",
        "Errors",
        "Processing Rate %",
        "Error Rate %"
    ]
    
    # Write headers (Row 1)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
    
    # Write data (Row 2)
    current_date = datetime.now().strftime("%Y-%m-%d")
    data_row = [
        current_date,
        daily_metrics.get('total_count', 0),
        daily_metrics.get('total_processed', 0),
        daily_metrics.get('total_errors', 0),
        daily_metrics.get('processing_rate', 0.0),
        daily_metrics.get('error_rate', 0.0)
    ]
    
    for col_num, value in enumerate(data_row, start=1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = value
        
        # Format percentage columns (columns 5 and 6)
        if col_num in [5, 6]:
            cell.number_format = '0.0"%"'
    
    # Apply formatting
    _format_sheet_headers(sheet)
    _auto_adjust_columns(sheet)


def _create_mtd_summary_sheet(sheet, mtd_metrics: dict) -> None:
    """
    Create and populate the MTD Summary sheet.
    
    Args:
        sheet: openpyxl worksheet object
        mtd_metrics (dict): Dictionary containing MTD metrics
    """
    # Define headers
    headers = [
        "Period",
        "Total Count",
        "Processed",
        "Errors",
        "Processing Rate %",
        "Error Rate %",
        "Days",
        "Daily Average"
    ]
    
    # Write headers (Row 1)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
    
    # Write data (Row 2)
    data_row = [
        "Month-to-Date",
        mtd_metrics.get('mtd_total', 0),
        mtd_metrics.get('mtd_processed', 0),
        mtd_metrics.get('mtd_errors', 0),
        mtd_metrics.get('processing_rate', 0.0),
        mtd_metrics.get('error_rate', 0.0),
        mtd_metrics.get('num_days', 0),
        mtd_metrics.get('daily_average', 0)
    ]
    
    for col_num, value in enumerate(data_row, start=1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = value
        
        # Format percentage columns (columns 5 and 6)
        if col_num in [5, 6]:
            cell.number_format = '0.0"%"'
    
    # Apply formatting
    _format_sheet_headers(sheet)
    _auto_adjust_columns(sheet)


def _create_warnings_sheet(sheet, validation_warnings: list) -> None:
    """
    Create and populate the Warnings sheet.
    
    Args:
        sheet: openpyxl worksheet object
        validation_warnings (list): List of warning messages
    """
    # Write header (Row 1)
    header_cell = sheet.cell(row=1, column=1)
    header_cell.value = "Validation Warnings"
    
    # Write each warning in subsequent rows
    for row_num, warning in enumerate(validation_warnings, start=2):
        cell = sheet.cell(row=row_num, column=1)
        cell.value = warning
    
    # Apply formatting
    _format_sheet_headers(sheet)
    _auto_adjust_columns(sheet)


def _format_sheet_headers(sheet) -> None:
    """
    Format the first row of a sheet with bold font.
    
    This is a helper function that makes headers stand out by applying
    bold formatting to all cells in the first row.
    
    Args:
        sheet: openpyxl worksheet object to format
    """
    # Create bold font style
    bold_font = Font(bold=True)
    
    # Apply bold font to all cells in the first row
    for cell in sheet[1]:
        cell.font = bold_font


def _auto_adjust_columns(sheet) -> None:
    """
    Auto-adjust column widths based on content.
    
    This helper function iterates through all columns and sets their width
    to accommodate the longest content in each column, with a small padding.
    
    Args:
        sheet: openpyxl worksheet object to adjust
    """
    # Iterate through all columns
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Find the maximum length of content in this column
        for cell in column:
            try:
                if cell.value:
                    # Convert to string and measure length
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width with some padding
        # Minimum width of 10, maximum of 50 for readability
        adjusted_width = min(max(max_length + 2, 10), 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Made with Bob
